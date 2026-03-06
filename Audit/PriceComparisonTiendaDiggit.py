import sys
import pandas as pd
from pathlib import Path
import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# ---------------------------------------------------------
# 📂 Forzar ctcVendorController al PYTHONPATH
# ---------------------------------------------------------
BASE_DIR = Path(__file__).resolve()
while BASE_DIR.name != "CTCVENDORCONTROLLER":
    BASE_DIR = BASE_DIR.parent

sys.path.insert(0, str(BASE_DIR))

DATA_DIR = BASE_DIR / "DataStorage"

# ---------------------------------------------------------
# 🧠 Vendor DIGGIT (desde ICBC)
# ---------------------------------------------------------
from config.icbc_vendors import ICBC_VENDORS

DIGGIT_VENDOR_KEY = "diggit"

if DIGGIT_VENDOR_KEY not in ICBC_VENDORS:
    raise KeyError("❌ No existe el vendor 'DIGGIT' en icbc_vendors.py")

DIGGIT_VENDOR_ID = str(ICBC_VENDORS[DIGGIT_VENDOR_KEY]["vendor_id"])

# ---------------------------------------------------------
# 🔍 Buscar el último archivo según patrón
# ---------------------------------------------------------
def get_latest_csv(pattern: str) -> Path:
    files = sorted(
        [f for f in DATA_DIR.glob(pattern)],
        key=lambda x: x.stat().st_mtime,
        reverse=True
    )
    if not files:
        raise FileNotFoundError(f"❌ No se encontró archivo con patrón {pattern}")

    print(f"📄 Archivo encontrado: {files[0].name}")
    return files[0]

# ---------------------------------------------------------
# 🧠 Auditoría TIENDA DIGGIT
# ---------------------------------------------------------
def audit_diggit():

    # -----------------------------------------------------
    # 📥 Productos Patagonia (precio Patagonia)
    # -----------------------------------------------------
    print("\n🔍 Cargando ProductosPatagonia_* (precio Patagonia)...")
    patagonia_path = get_latest_csv("ProductosPatagonia_*.csv")
    df_patagonia = pd.read_csv(patagonia_path, dtype=str)

    required_cols = ["sku", "ProviderId", "PrecioVenta"]
    for col in required_cols:
        if col not in df_patagonia.columns:
            raise KeyError(
                f"❌ No existe la columna {col} en {patagonia_path.name}"
            )

    # Filtrar solo DIGGIT
    df_diggit_prod = df_patagonia[
        df_patagonia["ProviderId"] == DIGGIT_VENDOR_ID
    ].copy()

    if df_diggit_prod.empty:
        raise ValueError(
            "❌ No se encontraron productos DIGGIT en ProductosPatagonia"
        )

    # Precios Patagonia
    df_diggit_prod["SKU_PATAGONIA"] = df_diggit_prod["sku"]
    df_diggit_prod["PRICE_PATAGONIA"] = df_diggit_prod["PrecioVenta"].astype(float)

    df_diggit_prod = df_diggit_prod[
        ["SKU_PATAGONIA", "PRICE_PATAGONIA"]
    ]

    # -----------------------------------------------------
    # 📥 Precios Web DIGGIT
    # -----------------------------------------------------
    print("\n🔍 Cargando Diggit_Precios_* ...")
    diggit_price_path = get_latest_csv("Diggit_Precios_*.csv")
    df_diggit_price = pd.read_csv(diggit_price_path, dtype=str)

    if "PRICE_DIGGIT" not in df_diggit_price.columns:
        raise KeyError(
            f"❌ No existe la columna PRICE_DIGGIT en {diggit_price_path.name}"
        )

    df_diggit_price["PRICE_DIGGIT"] = df_diggit_price["PRICE_DIGGIT"].astype(float)

    # -----------------------------------------------------
    # 🔗 Merge por SKU_PATAGONIA
    # -----------------------------------------------------
    df_final = pd.merge(
        df_diggit_prod,
        df_diggit_price[["SKU_PATAGONIA", "PRICE_DIGGIT"]],
        on="SKU_PATAGONIA",
        how="left"
    )

    # -----------------------------------------------------
    # 🧮 Cálculo de diferencias
    # -----------------------------------------------------
    df_final["DIF_ABSOLUTA"] = (
        df_final["PRICE_PATAGONIA"] - df_final["PRICE_DIGGIT"]
    )
    df_final["DIF_PORCENTUAL"] = (
        (df_final["PRICE_PATAGONIA"] / df_final["PRICE_DIGGIT"]) - 1
    )

    # -----------------------------------------------------
    # 🎯 Formato final
    # -----------------------------------------------------
    df_final = df_final[
        [
            "SKU_PATAGONIA",
            "PRICE_PATAGONIA",
            "PRICE_DIGGIT",
            "DIF_ABSOLUTA",
            "DIF_PORCENTUAL"
        ]
    ]

    print("\n✅ Auditoría Tienda Diggit completada.\n")
    return df_final

# ---------------------------------------------------------
# 🎨 Guardar Excel con colores
# ---------------------------------------------------------
def save_excel_with_colors(df: pd.DataFrame):
    print("🎨 Generando Excel con colores ...")

    wb = Workbook()
    ws = wb.active
    ws.title = "Auditoria Diggit"

    headers = list(df.columns)
    ws.append(headers)

    RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    for _, row in df.iterrows():
        ws.append(list(row.values))
        excel_row = ws.max_row

        pct = row["DIF_PORCENTUAL"]

        if pct >= 0.10:
            fill = RED      # Patagonia ≥ 10% más caro
        elif pct < 0:
            fill = GREEN    # Patagonia más barato
        else:
            fill = YELLOW   # Entre 0% y 10%

        for col in range(1, len(headers) + 1):
            ws.cell(row=excel_row, column=col).fill = fill

    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = DATA_DIR / f"Audit_Diggit_{timestamp}.xlsx"
    wb.save(out_path)

    print(f"\n💾 Excel guardado en:\n{out_path}\n")

# ---------------------------------------------------------
# ▶ Run
# ---------------------------------------------------------
if __name__ == "__main__":
    df_result = audit_diggit()
    save_excel_with_colors(df_result)

