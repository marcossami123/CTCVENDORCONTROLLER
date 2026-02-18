import os
import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ---------------------------------------------------------
# Obtener último archivo por prefijo
# ---------------------------------------------------------
def get_latest_file(data_storage_path: Path, prefix: str) -> Path:
    csv_files = [f for f in data_storage_path.glob("*.csv") if prefix in f.name]
    if not csv_files:
        raise FileNotFoundError(
            f"No se encontraron archivos con prefijo '{prefix}' en {data_storage_path}"
        )

    latest_file = max(csv_files, key=os.path.getmtime)
    print(f"Ultimo archivo encontrado para '{prefix}': {latest_file.name}")
    return latest_file

# ---------------------------------------------------------
# Normalización de SKU
# ---------------------------------------------------------
def normalize_sku(s):
    if pd.isna(s):
        return ""

    s = str(s)

    # Quitar notación científica
    if "e+" in s.lower():
        try:
            s = "{:.0f}".format(float(s))
        except Exception:
            pass

    # Eliminar caracteres invisibles
    for char in ["\ufeff", "\u200b", "\u00a0"]:
        s = s.replace(char, "")

    return s.strip().upper().replace(" ", "")

# ---------------------------------------------------------
# MAIN
# ---------------------------------------------------------
def main():
    base_path = Path(__file__).resolve().parent.parent
    data_storage_path = base_path / "DataStorage"

    # Archivos de entrada
    patagonia_file = get_latest_file(data_storage_path, "ProductosPatagonia")
    megatone_file = get_latest_file(data_storage_path, "megatone_prices")

    df_patagonia = pd.read_csv(patagonia_file, dtype=str)
    df_megatone = pd.read_csv(megatone_file, dtype=str)

    # Columnas
    col_sku_pat = "sku"
    col_precio_pat = "PrecioVenta"

    col_sku_mega = "SKU_Patagonia"
    col_precio_mega = "Sale_Price"

    # Validaciones
    for col in [col_sku_pat, col_precio_pat]:
        if col not in df_patagonia.columns:
            raise ValueError(f"La columna '{col}' no está en Patagonia")

    for col in [col_sku_mega, col_precio_mega]:
        if col not in df_megatone.columns:
            raise ValueError(f"La columna '{col}' no está en Megatone")

    # Normalizar SKUs
    df_patagonia["SKU_norm"] = df_patagonia[col_sku_pat].apply(normalize_sku)
    df_megatone["SKU_norm"] = df_megatone[col_sku_mega].apply(normalize_sku)

    # Convertir precios
    df_patagonia["PrecioVenta"] = pd.to_numeric(
        df_patagonia["PrecioVenta"], errors="coerce"
    )
    df_megatone["Sale_Price"] = pd.to_numeric(
        df_megatone["Sale_Price"], errors="coerce"
    )

    # Merge
    df_auditoria = pd.merge(
        df_patagonia[[col_sku_pat, "SKU_norm", "PrecioVenta"]],
        df_megatone[[col_sku_mega, "SKU_norm", "Sale_Price"]],
        on="SKU_norm",
        how="left"
    )

    # Cálculos
    df_auditoria["Diferencia_Absoluta"] = (
        df_auditoria["PrecioVenta"] - df_auditoria["Sale_Price"]
    )
    df_auditoria["Diferencia_Porcentual"] = (
        df_auditoria["PrecioVenta"] / df_auditoria["Sale_Price"]
    ) - 1

    # Columnas finales
    df_auditoria = df_auditoria.rename(columns={col_sku_pat: "SKU"})
    df_auditoria = df_auditoria[
        [
            "SKU",
            "PrecioVenta",
            "Sale_Price",
            "Diferencia_Absoluta",
            "Diferencia_Porcentual",
        ]
    ]

    # Exportar CSV
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_file = data_storage_path / f"ResultadoDeAuditoria_{timestamp}.csv"
    df_auditoria.to_csv(csv_file, index=False, encoding="utf-8-sig")

    print("Auditoria CSV generada correctamente:", csv_file)

    # Exportar Excel
    excel_file = data_storage_path / f"Audit_Megatone_{timestamp}.xlsx"
    df_auditoria.to_excel(excel_file, index=False)

    # Formato condicional Excel
    wb = load_workbook(excel_file)
    ws = wb.active

    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    col_index = list(df_auditoria.columns).index("Diferencia_Porcentual") + 1

    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=col_index).value
        if val is None:
            continue

        fill = red if val > 0.10 else green if val < 0 else yellow
        for c in ws[row]:
            c.fill = fill

    wb.save(excel_file)

    print("Archivo Excel generado correctamente con formato:", excel_file)
    print(df_auditoria.head())

# ---------------------------------------------------------
# Entry point
# ---------------------------------------------------------
if __name__ == "__main__":
    main()












