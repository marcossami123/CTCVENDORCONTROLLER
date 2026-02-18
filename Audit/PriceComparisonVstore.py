import sys
import pandas as pd
from pathlib import Path
import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# ---------------------------------------------------------
# 📂 Forzar ctcVendorController al PYTHONPATH
# ---------------------------------------------------------
BASE_DIR = Path(__file__).resolve()
while BASE_DIR.name != "ctcVendorController":
    BASE_DIR = BASE_DIR.parent

sys.path.insert(0, str(BASE_DIR))

DATA_DIR = BASE_DIR / "DataStorage"

# ---------------------------------------------------------
# 🧠 Vendor VSTORE (desde ICBC)
# ---------------------------------------------------------
from config.icbc_vendors import ICBC_VENDORS

VSTORE_VENDOR_KEY = "vstore"

if VSTORE_VENDOR_KEY not in ICBC_VENDORS:
    raise KeyError("❌ No existe el vendor 'VSTORE' en icbc_vendors.py")

VSTORE_VENDOR_ID = str(ICBC_VENDORS[VSTORE_VENDOR_KEY]["vendor_id"])

# ---------------------------------------------------------
# 🔍 Buscar el último archivo según patrón
# ---------------------------------------------------------
def get_latest_csv(pattern: str) -> Path:
    files = sorted(
        [f for f in DATA_DIR.glob(pattern)],
        key=lambda x: x.stat().st_mtime,
        reverse=True
    )
    if not files:
        raise FileNotFoundError(f"❌ No se encontró archivo con patrón {pattern}")
    print(f"📄 Archivo encontrado: {files[0].name}")
    return files[0]

# ---------------------------------------------------------
# 🧠 Auditoría VSTORE
# ---------------------------------------------------------
def audit_vstore():

    # -----------------------------------------------------
    # 📥 Productos Patagonia (precio Patagonia)
    # -----------------------------------------------------
    print("\n🔍 Cargando ProductosPatagonia_* (precio Patagonia)...")
    patagonia_path = get_latest_csv("ProductosPatagonia_*.csv")
    df_patagonia = pd.read_csv(patagonia_path, dtype=str)

    required_cols = ["sku", "ProviderId", "PrecioVenta"]
    for col in required_cols:
        if col not in df_patagonia.columns:
            raise KeyError(
                f"❌ No existe la columna {col} en {patagonia_path.name}"
            )

    # Filtrar solo VSTORE
    df_vstore_prod = df_patagonia[
        df_patagonia["ProviderId"] == VSTORE_VENDOR_ID
    ].copy()

    if df_vstore_prod.empty:
        raise ValueError(
            "❌ No se encontraron productos VSTORE en ProductosPatagonia"
        )

    # Precios Patagonia
    df_vstore_prod["SKU_PATAGONIA"] = df_vstore_prod["sku"]
    df_vstore_prod["PRICE_PATAGONIA"] = df_vstore_prod["PrecioVenta"].astype(float)

    df_vstore_prod = df_vstore_prod[
        ["SKU_PATAGONIA", "PRICE_PATAGONIA"]
    ]

    # -----------------------------------------------------
    # 📥 Precios Web VSTORE
    # -----------------------------------------------------
    print("\n🔍 Cargando Vstore_Precios_* ...")
    vstore_price_path = get_latest_csv("Vstore_Precios_*.csv")
    df_vstore_price = pd.read_csv(vstore_price_path, dtype=str)

    if "PRICE_VSTORE" not in df_vstore_price.columns:
        raise KeyError(
            f"❌ No existe la columna PRICE_VSTORE en {vstore_price_path.name}"
        )

    df_vstore_price["PRICE_VSTORE"] = df_vstore_price["PRICE_VSTORE"].astype(float)

    # -----------------------------------------------------
    # 🔗 Merge por SKU_PATAGONIA
    # -----------------------------------------------------
    df_final = pd.merge(
        df_vstore_prod,
        df_vstore_price[["SKU_PATAGONIA", "PRICE_VSTORE"]],
        on="SKU_PATAGONIA",
        how="left"
    )

    # -----------------------------------------------------
    # 🧮 Calcular diferencias
    # -----------------------------------------------------
    df_final["DIF_ABSOLUTA"] = (
        df_final["PRICE_PATAGONIA"] - df_final["PRICE_VSTORE"]
    )
    df_final["DIF_PORCENTUAL"] = (
        (df_final["PRICE_PATAGONIA"] / df_final["PRICE_VSTORE"]) - 1
    )

    # -----------------------------------------------------
    # 🎯 Formato final
    # -----------------------------------------------------
    df_final = df_final[
        [
            "SKU_PATAGONIA",
            "PRICE_PATAGONIA",
            "PRICE_VSTORE",
            "DIF_ABSOLUTA",
            "DIF_PORCENTUAL"
        ]
    ]

    print("\n✅ Auditoría Vstore completada.\n")
    return df_final

# ---------------------------------------------------------
# 🎨 Guardar Excel con colores
# ---------------------------------------------------------
def save_excel_with_colors(df: pd.DataFrame):
    print("🎨 Generando Excel con colores ...")

    wb = Workbook()
    ws = wb.active
    ws.title = "Auditoria Vstore"

    headers = list(df.columns)
    ws.append(headers)

    RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    for _, row in df.iterrows():
        ws.append(list(row.values))
        excel_row = ws.max_row

        pct = row["DIF_PORCENTUAL"]

        if pct >= 0.10:
            fill = RED      # Patagonia ≥ 10% más caro
        elif pct < 0:
            fill = GREEN    # Patagonia más barato
        else:
            fill = YELLOW   # Entre 0% y 10%

        for col in range(1, len(headers) + 1):
            ws.cell(row=excel_row, column=col).fill = fill

    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = DATA_DIR / f"Audit_Vstore_{timestamp}.xlsx"
    wb.save(out_path)

    print(f"\n💾 Excel guardado en:\n{out_path}\n")

# ---------------------------------------------------------
# ▶ Run
# ---------------------------------------------------------
if __name__ == "__main__":
    df_result = audit_vstore()
    save_excel_with_colors(df_result)


