import sys
import os
import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ---------------------------------------------------------
# 📂 Forzar ctcVendorController al PYTHONPATH
# ---------------------------------------------------------
BASE_DIR = Path(__file__).resolve()
while BASE_DIR.name != "CTCVENDORCONTROLLER":
    BASE_DIR = BASE_DIR.parent

sys.path.insert(0, str(BASE_DIR))

DATA_DIR = BASE_DIR / "DataStorage"

# ---------------------------------------------------------
# 🧠 Vendor Radio Sapienza (desde ICBC)
# ---------------------------------------------------------
from config.icbc_vendors import ICBC_VENDORS

SAPIENZA_VENDOR_KEY = "radio sapienza"

if SAPIENZA_VENDOR_KEY not in ICBC_VENDORS:
    raise KeyError("❌ No existe el vendor 'radio sapienza' en icbc_vendors.py")

SAPIENZA_VENDOR_ID = str(ICBC_VENDORS[SAPIENZA_VENDOR_KEY]["vendor_id"])

# ---------------------------------------------------------
# 📄 Obtener último archivo por prefijo
# ---------------------------------------------------------
def get_latest_file(data_storage_path: Path, prefix: str) -> Path:
    csv_files = [
        f for f in data_storage_path.glob("*.csv")
        if f.name.startswith(prefix)
    ]
    if not csv_files:
        raise FileNotFoundError(
            f"No se encontraron archivos con prefijo '{prefix}' en {data_storage_path}"
        )

    latest_file = max(csv_files, key=os.path.getmtime)
    print(f"📄 Último archivo encontrado para '{prefix}': {latest_file.name}")
    return latest_file

# ---------------------------------------------------------
# 🔎 Normalización de SKU
# ---------------------------------------------------------
def normalize_sku(s):
    if pd.isna(s):
        return ""

    s = str(s)

    # Quitar notación científica
    if "e+" in s.lower():
        try:
            s = "{:.0f}".format(float(s))
        except Exception:
            pass

    # Eliminar caracteres invisibles
    for char in ["\ufeff", "\u200b", "\u00a0"]:
        s = s.replace(char, "")

    return s.strip().upper().replace(" ", "")

# ---------------------------------------------------------
# ▶️ MAIN
# ---------------------------------------------------------
def main():

    # -----------------------------------------------------
    # 1️⃣ Productos Patagonia (precio Patagonia)
    # -----------------------------------------------------
    print("\n🔍 Cargando ProductosPatagonia_* (precio Patagonia)...")
    patagonia_file = get_latest_file(DATA_DIR, "ProductosPatagonia")

    df_patagonia = pd.read_csv(patagonia_file, dtype=str)

    required_cols = ["sku", "ProviderId", "PrecioVenta"]
    for col in required_cols:
        if col not in df_patagonia.columns:
            raise ValueError(
                f"❌ La columna '{col}' no está en {patagonia_file.name}"
            )

    # Filtrar solo Radio Sapienza
    df_sapienza = df_patagonia[
        df_patagonia["ProviderId"] == SAPIENZA_VENDOR_ID
    ].copy()

    if df_sapienza.empty:
        raise ValueError(
            "❌ No se encontraron productos Radio Sapienza en ProductosPatagonia"
        )

    # -----------------------------------------------------
    # 2️⃣ Precios Web / API Radio Sapienza
    # -----------------------------------------------------
    prices_file = get_latest_file(DATA_DIR, "radiosapienza_prices")
    df_prices = pd.read_csv(prices_file, dtype=str)

    col_sku_prices = "SKUdb"
    col_precio_prices = "precio"

    for col in [col_sku_prices, col_precio_prices]:
        if col not in df_prices.columns:
            raise ValueError(
                f"❌ La columna '{col}' no está en {prices_file.name}"
            )

    # -----------------------------------------------------
    # 3️⃣ Normalizar SKUs
    # -----------------------------------------------------
    df_sapienza["SKU_norm"] = df_sapienza["sku"].apply(normalize_sku)
    df_prices["SKU_norm"] = df_prices[col_sku_prices].apply(normalize_sku)

    # -----------------------------------------------------
    # 4️⃣ Convertir precios a numérico
    # -----------------------------------------------------
    df_sapienza["PrecioVenta"] = pd.to_numeric(
        df_sapienza["PrecioVenta"], errors="coerce"
    )
    df_prices["PrecioAPI"] = pd.to_numeric(
        df_prices[col_precio_prices], errors="coerce"
    )

    # -----------------------------------------------------
    # 5️⃣ Merge
    # -----------------------------------------------------
    df_auditoria = pd.merge(
        df_sapienza[["sku", "SKU_norm", "PrecioVenta"]],
        df_prices[[col_sku_prices, "SKU_norm", "PrecioAPI"]],
        on="SKU_norm",
        how="left"
    )

    # -----------------------------------------------------
    # 6️⃣ Cálculos
    # -----------------------------------------------------
    df_auditoria["Diferencia_Absoluta"] = (
        df_auditoria["PrecioVenta"] - df_auditoria["PrecioAPI"]
    )

    df_auditoria["Diferencia_Porcentual"] = (
        (df_auditoria["PrecioVenta"] / df_auditoria["PrecioAPI"]) - 1
    )

    # -----------------------------------------------------
    # 7️⃣ Columnas finales
    # -----------------------------------------------------
    df_auditoria = df_auditoria.rename(columns={"sku": "SKU"})
    df_auditoria = df_auditoria[
        [
            "SKU",
            "PrecioVenta",
            "PrecioAPI",
            "Diferencia_Absoluta",
            "Diferencia_Porcentual",
        ]
    ]

    # -----------------------------------------------------
    # 8️⃣ Exportar CSV
    # -----------------------------------------------------
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_file = DATA_DIR / f"ResultadoDeAuditoria_RadioSapienza_{timestamp}.csv"
    df_auditoria.to_csv(csv_file, index=False, encoding="utf-8-sig")

    print("\n✅ Auditoría CSV generada correctamente:", csv_file)

    # -----------------------------------------------------
    # 9️⃣ Exportar Excel
    # -----------------------------------------------------
    excel_file = DATA_DIR / f"Audit_RadioSapienza_{timestamp}.xlsx"
    df_auditoria.to_excel(excel_file, index=False)

    # -----------------------------------------------------
    # 🔟 Formato condicional Excel
    # -----------------------------------------------------
    wb = load_workbook(excel_file)
    ws = wb.active

    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    col_index = list(df_auditoria.columns).index("Diferencia_Porcentual") + 1

    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=col_index).value
        if val is None:
            continue

        fill = red if val > 0.10 else green if val < 0 else yellow
        for c in ws[row]:
            c.fill = fill

    wb.save(excel_file)

    print(f"📊 Archivo Excel generado correctamente con formato: {excel_file}")
    print(df_auditoria.head())

# ---------------------------------------------------------
# ▶️ Entry point
# ---------------------------------------------------------
if __name__ == "__main__":
    main()

